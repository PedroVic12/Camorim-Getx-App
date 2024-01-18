import openpyxl
import pyautogui
import pandas as pd
PLANILHA = '../assets/EQUIPAMENTOS FROTA (1).xlsx'


def escreve_formulario(x,y,dados):
    pyautogui.click(x,y, duration = 0.5)
    pyautogui.write(dados)

def get_last_row(planilha):
    return planilha.max_row

def get_index(workbook):
    for index, sheet in enumerate(workbook, start=1):
        sheet.title = f'Aba{index}'


def get_cell(sheet,range):
    sheet[range].value 
    
def ver_planilha(wk_name):
    for row in wk_name.iter_rows(values_only=True):
        print(row)

def preenche_dados(sheet,df,index):
    for row in df.iterrows():
        sheet[f'B{9 + index}'] = row['EQUIPAMENTO']
        sheet[f'D{9 + index}'] = row['FABRICANTE']
        sheet[f'E{9 + index}'] = row['MODELO']



def salvar(wb,name_xlsx):
    wb.save(f'{name_xlsx}.xlsx')
    print(f'Arquivo {name_xlsx} salvo!')

def copiar_aba(wb,nome_old,nome_new):
    sheet_to_copy = wb[nome_old]
    new_sheet = wb.create_sheet(nome_new)
    for row in sheet_to_copy.iter_rows():
        for cell in row:
            #print(cell)
            new_sheet[cell.coordinate].value = cell.value

def create_sheets(wb,array):
    for i in array:
        wb.create_sheet(title= i)
        wb.save('../assets/Cadastro de Equipamentos - Oilxplorer Marine.xlsx')

        
def preencher_aba(workbook, aba, df, nome_barco, start_row):
    # Filtrar os dados para a embarcação atual
    df_filtrado = df.loc[df['EMBARCAÇÃO'] == nome_barco]
    aba['D3'].value = nome_barco

    for i, row in enumerate(df_filtrado.itertuples(), start=start_row):
        print(df_filtrado)
        aba[f'B{i}'].value = row.EQUIPAMENTO
        aba[f'D{i}'].value = row.FABRICANTE
        aba[f'E{i}'].value = row.MODELO

def copiar_excel(wb,new_sheet_name):
    source = wb.active
    target = wb.copy_worksheet(source)
    target.title = new_sheet_name
    
    salvar(wb,'copia_teste')

def roboPreencheCadastro(PLANILHA):

    workbook = openpyxl.load_workbook(PLANILHA)
    planilha = workbook['RESUMO']
    _array = []

    lastRow = get_last_row(planilha)

    for row in planilha.iter_rows(min_row = 2,max_row=5):

        # COluna1 , Coluna 2, Coluna 3
        equipamento = row[1].value
        print(equipamento)
        escreve_formulario(1150,235,equipamento)

        rebocador = row[0].value
        escreve_formulario(1150,315,rebocador)

        acao = row[2].value
        escreve_formulario(1150,385,acao)

        status = row[3].value
        escreve_formulario(1150,465,status)

        tipo = row[4].value
        escreve_formulario(1150,515,tipo)
    
        pyautogui.click(1319,724)

PLANILHA = '../assets/EQUIPAMENTOS FROTA (1).xlsx'

roboPreencheCadastro(PLANILHA)